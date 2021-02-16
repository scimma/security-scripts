"""
Provide base classes for acquisition of data and for  measurement/analysis.

The Database class loads information into a sqlite relational DB.
The Measurement class uses the relational db.
The two classes do not otherwise communicate.
"""
from security_scripts.information.lib import shlog
import pandas as pd
import fnmatch
import boto3
from security_scripts.information.lib import vanilla_utils
import json
import os
from security_scripts.information.lib import aws_utils
import glob
import pyjq
import botocore
from collections import ChainMap
from itertools import product


universal_services = ["sts", "iam", "route53", "route53domains", "s3", "s3control", "cloudfront", "organizations"]


class Dataset:
    """
    Base Class for  acquiring and  preparing a data set of interest for an analysis.
    """
    
    are_shared_tables_made = False
    
    def __init__(self, args, name, q):
        self.args=args
        self.q=q
        self.name=name
        self.table_name = None
        # dir enforcement
        self.args.report_path = os.getcwd() + "/report_files"
        if not os.path.exists(self.args.report_path):
            os.makedirs(self.args.report_path)

    def _insert_all_json(self, resource_name, id, record):
        """
        Load record into standard JSON table.

        The  all_json table allows scanning for untagged items.
        """
        if not Dataset.are_shared_tables_made:        
            sql = "CREATE TABLE all_json (resource_name TEXT, id TEXT, record JSON)"
            self.q.q(sql)
            Dataset.are_shared_tables_made = True
        sql="INSERT INTO all_json VALUES (?,?,?)"
        ilist = (resource_name, id, record)
        self.q.executemany(sql,[ilist])

 
    def does_table_exist(self):
        """
        Determine if table exist, indicating it has been populated w/ cached data
        """
        import sqlite3 #for exception value
        try:
            sql = "SELECT * FROM {}".format(self.table_name)
            self.q.q(sql)
            return True
        except sqlite3.OperationalError:
            return False
        
    def _json_clean_dumps(self, json_native): # L0B?
        """ convert native json into text

             Deal with Amazon stuff that is awkward in
             thw tools  set(python, etc)
        """

        # Amazon binarties have datetime types
        # need a user supplined encoder method
        # to turn to text I've not looked  into tiemzones.
        enc = vanilla_utils.DateTimeEncoder  
        json_text = json.dumps(json_native, cls=enc)
        #
        # I found no good way to deal with the "arrays of
        # dicionaary with basically the same keys  (like
        # how AWS present tags)
        # WOudl love some code here
        return json_text

    def get_page_iterator(self, aws_client_name, aws_function_name, region_name, parameter=None):
        client = self.args.session.client(aws_client_name, region_name=region_name)
        paginator = client.get_paginator(aws_function_name)
        shlog.verbose('about to iterate: {} {}'.format(aws_client_name, aws_function_name))
        if parameter is not None:  # more stable than straight up **parameters
            page_iterator = paginator.paginate(**parameter)
        else:
            page_iterator = paginator.paginate()
        return page_iterator

    def get_unpaginatable_data(self, client, aws_function_name, parameter=None):
        function = getattr(client, aws_function_name)
        try:
            if parameter is not None:  # more stable than straight up **parameters
                data = function(**parameter)
                for p in parameter:
                    data[p] = parameter[p]
            else:
                data = function()
        except Exception as e:
            if "ResourceNotFoundException" in str(e) \
                    or "NoSuchBucketPolicy" in str(e)\
                    or "ServerSideEncryptionConfigurationNotFoundError" in str(e):
                # happens when the resource exists, but in a different region
                print("resource not found!")
                return {'Nothing':[], 'ResponseMetadata':{}}  # this simulates empty output
            elif "InvalidParameter" in str(e):
                print("InvalidParameterException thrown, ignoring...")
                return {'Nothing': [], 'ResponseMetadata': {}}  # this simulates empty output
            elif "NotFound" in str(e) or "NoSuchEntity" in str(e):
                print("resource not found!")
                return {'Nothing': [], 'ResponseMetadata': {}}  # this simulates empty output
            elif "NoSuchTagSet" in str(e):
                print("no tags for resource!")
                return {'Nothing': [], 'ResponseMetadata': {}}  # this simulates empty output
            else:
                # still want to raise
                raise e
        return data


    def kwjq(self, recipes):
        """generate extra call arguments using jq
        if there's nothing to generate, cop out with output=json that will result in a single run"""
        parameters = []
        for recipe in recipes:
            temp = []
            if ".json" in recipes[recipe]:
                # break down string into:
                # parameter
                param = recipe
                # filename
                sp = recipes[recipe].split("|", 1)
                filename = sp[0]
                # jq string
                jq_str = sp[1]
                # ingest file from L0A/filename
                ref_file = self.json_from_file(self.f_path + filename) # l0a
                # let jq loose on the ingested file
                findings = pyjq.all(jq_str, ref_file)
                # append jq output to parameters
                for finding in findings:
                    temp.append({param:finding})
            else:
                # TODO: static, non jq handling
                temp.append({recipe:recipes[recipe]})
            parameters.append(temp)
        return [dict(ChainMap(*a)) for a in list(product(*parameters))]

    def page_param_setter(self, aws_client_name, aws_function_name, recipes):
        # run _pages_all_regions as many times as needed
        if recipes == {}:
            yield from self._pages_all_regions(aws_client_name, aws_function_name)
        else:
            parameters = self.kwjq(recipes)
            for parameter in parameters:
                yield from self._pages_all_regions(aws_client_name, aws_function_name, parameter=parameter)

    def _pages_all_regions(self, aws_client_name, aws_function_name, parameter=None):
        """
        An interator that gets the pages and boto client for all regions.

        aws_client_name = "resorurcemappingapi"
        functon_name = "get_resources"

        Can be used in for loops producing list of pages on a topic
        """
        self.args.session = boto3.Session(profile_name=self.args.profile)
        region_name_list = aws_utils.decribe_regions_df(self.args)['RegionName']

        # init a client to check for pagination
        client = self.args.session.client(aws_client_name, region_name=region_name_list[0])

        if client.can_paginate(aws_function_name):
            if aws_client_name in universal_services:
                page_iterator = self.get_page_iterator(aws_client_name, aws_function_name, region_name_list[0], parameter)
                for page in page_iterator:
                    # augment if parameterized
                    if parameter is not None:
                        for p in parameter:
                            page[p] = parameter[p]
                    yield page, None
            else:
                for region_name in region_name_list:
                    page_iterator = self.get_page_iterator(aws_client_name, aws_function_name, region_name, parameter)
                    try:
                        for page in page_iterator:
                            if parameter is not None:
                                for p in parameter:
                                    page[p] = parameter[p]
                            yield page, None
                    except Exception as e:
                        if "ListPlatformApplications" in str(e) and "is not supported in this region" in str(e):
                            print("operation not supported in region, skipping...")
                            yield {'Nothing': [], 'ResponseMetadata': {}}, None
        else:
            if aws_client_name in universal_services:
                # use client initiated earlier
                yield self.get_unpaginatable_data(client, aws_function_name, parameter), None
            else:
                # loop through regions
                for region_name in region_name_list:
                    client = self.args.session.client(aws_client_name, region_name=region_name)
                    yield self.get_unpaginatable_data(client, aws_function_name, parameter), None
                    # TODO: get_unpaginatable_data gets a single kwargs entry
                    # loop through every kwarg possible
                    # 1. kwargs are collected elsewhere
                    # (this function returns pages with results)
                    # 2. get_unpaginatable_data gets **kwargs={param:arn}
                    # => _pages_all_regions needs to be run separately for each kwarg possible


    def json_from_file(self, filename):
        "return binary json contents from a file)"
        jf = open(filename, "r")
        jlist = json.load(jf)
        jf.close()
        return jlist

    def jsons_from_dir(self, dir):
        """
        return binary json contents of file
        and base file name.
        """
        file_glob = os.path.join(dir, "*.json")
        for filename in glob.iglob(file_glob):
            jlist = self.json_from_file(filename)
            yield jlist, os.path.basename(filename)

    def make_data(self):
        """ Build base of data needed for the indicated measurement """
        pass

    def clean_data(self):
        """  Remove items that are not relevant"""
        pass

    def print_data(self):
        """
        Print the dataset
        
        This method is provided primariy for debug.
        This method is useful after the make_data and clean steps
        """
        sql = "select * from %s" % (self.table_name)
        shlog.verbose(sql)
        df = self.q.q_to_df(sql)
        print(wrapped_ascii_table(self.args, df))

        
class Measurement:
    """
    Perform a number of tests against the data
    test methods begin with the characters tst_,
    and information methods beginning with inf_
        
    The result of each test is a dataframe
    giving information about the defects found.
    An Empty data frame indicates that no defects were
    found.

    The Result of each informational method is
    a tabular presentation on the screen.
    """
    
    def __init__(self, args, name, q):
        self.args=args
        self.q=q
        self.name=name
        self.df = None
        self.current_test=None
        self.listonly = args.listonly


        if self.listonly:
            self._print_tests("tst_")
            self._print_tests("inf_")
            self._print_tests("make_")
            self._print_tests("json_")
        else:
            self._call_analysis_methods("tst_",self._write_relational_files)
            self._call_analysis_methods("inf_",self._write_relational_files)
            self._call_analysis_methods("make_",self._call_asset_analyis)
            self._call_analysis_methods("json_",self._write_json_file)

    def _call_analysis_methods(self,prefix, report_func):
        """
        Call all methods begining with indicated prefix.
        
        This code uses the python metaclass system
        to find all methods beginning with the indiccated
        prefix and  invoke them.  These methods must have
        no arguments (other than self)

        After testing the report function supplied by the caller
        is invoked.
        """
        

        for name, func in self._list_tests(prefix):
            shlog.normal("starting analysis: %s" % (name))
            #result = func()
            report_func(func)


    def _list_tests(self, prefix):
        """
        return list of all tests

        each list element is a parit of  [ascii_name, function-to-call]
        filter according to th eglob in args.only
        """
        list = []
        for key in dir(self):
            if key[0:len(prefix)] == prefix:
                if not fnmatch.fnmatch(key, self.args.only) : continue
                list.append([key, self.__getattribute__(key)])
        return list

    #
    #  Here are routines that render the output
    #
    
    def _call_asset_analyis(self, func) :
        """
        Make the input table for asseet analysis.
        
        """
        func()

    def _print_tests(self, result, prefix):
        """
        Print a list of available tests with the indicated prefix

        Print to stdout.
        """
        for test, _ in self._list_tests(prefix):
            print(test)
        

    def _is_violation_detected(self):
        """Indicate that a test has not suceeded """
        if self.df.size == 0 :
            return False
        return True

    
    def _write_json_file (self, func):
        """
        Write a json file from a query containing one columm returned by func.

        The  file is useable by utilities such as jq.
        The file name is deried from  the name of the function.
        """
        name = func.__name__
        filename =  name + ".json"
        filepath = os.path.join(self.args.report_path, filename)
        shlog.normal ("preparing {}".format(filepath))
        sql = func()
        shlog.verbose("sql is: {}".format(sql))

        jsons = self.q.q(sql).fetchall()
        jsons = [j[0] for j in jsons]
        jsons = ",\n".join(jsons)
        jsons = "[" + jsons + "]"
        jsons = jsons.encode(encoding='UTF-8')
        # now  write it out.
        f = open(filepath,'wb')
        f.write(jsons)
        f.close()
        return

    def append_df_to_excel(self, filepath, df, sheet_name):
        """Append a DataFrame [df] to existing Excel file [filepath] into [sheet_name] Sheet.
        If [filepath] doesn't exist, then this function will create it.

        filepath: File path or existing ExcelWriter (Example: '/path/to/file.xlsx')
        df: dataframe to save to workbook
        sheet_name: Name of sheet which will contain DataFrame.
        startrow: upper left cell row to dump data frame.
                   Per default (startrow=None) calculate the last row
                   in the existing DF and write to the next row...
        truncate_sheet: truncate (remove and recreate) [sheet_name]
                         before writing DataFrame to Excel file
        to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`

        original https://stackoverflow.com/a/47740262/10086137
        """
        from openpyxl import load_workbook
        writer = pd.ExcelWriter(filepath, engine='openpyxl')

        # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError = IOError

        try:
            # try to open an existing workbook
            writer.book = load_workbook(filepath)
            # nuke sheet
            if sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            shlog.normal("creating excel file  {}".format(filepath))

        # write out the new sheet and save
        df.to_excel(writer, sheet_name, index=False)
        writer.save()

    def _write_relational_files(self, func):
        """
        Write a relational report out in tablular and csv formnt
        """
        name = func.__name__
        txt_filename =  name + ".txt"
        filepath = os.path.join(self.args.report_path, txt_filename)
        shlog.normal ("preparing {}".format(filepath))
        sql = func()
        shlog.verbose("sql is: {}".format(sql))
        df = self.q.q_to_df(sql)
        # wrapped table )qicj look for human
        table = wrapped_ascii_table(self.args, df).encode(encoding='UTF-8')
        f = open(filepath, "wb")
        f.write(table)
        f.close()
        # csv
        csv_filename =  name + ".csv"
        filepath = os.path.join(self.args.report_path, csv_filename)
        shlog.normal ("preparing {}".format(filepath))
        df.to_csv(filepath, index=False)
        # add to excel workbook
        # Thsi code is not working, it's be nice to have all this in workbook.
        # needs to be cloasds
        xlxs_filename =  "all.xlsx"
        filepath = os.path.join(self.args.report_path, xlxs_filename)
        # import pdb; pdb.set_trace()
        shlog.normal("adding sheet {} to {}".format(name, filepath))
        self.append_df_to_excel(filepath, df, name)

    
             
        
def wrapped_ascii_table(args, df):
    """
    Return a table with long lines wrapped within a table cell, given a df.

    The table is formed by tabulate 
    """
    import tabulate
    scratch = pd.DataFrame()  # do not alter caller's data
    if args.bare:
        table =tabulate.tabulate(df, list(df.columns))
        return table
    width = int(132/len(df.columns))
    for c in list(df.columns):
       #put newlines in the data frame
        scratch[c] = df[c].str.wrap(width)

       #Make a tabulate table, with wrapped data, column headnign and
       #ASCII art to draw cell boindaris.
    table =tabulate.tabulate(scratch, list(scratch.columns),"grid")
    return table   
